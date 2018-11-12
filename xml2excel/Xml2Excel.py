#!/usr/bin/env python
# -*- coding:utf-8 -*-
"""
@Author : yangwei.li
@Create date : 11/6/2018
@FileName : Xml2Excel.py
"""
import os
import re
import shutil
import openpyxl
from lxml import etree
from openpyxl.styles import Alignment,Font

# html标签
HTML_CHAR = ["<div>", "</div>", "<p>", "</p>", "&nbsp;"]

class Xml2Excel:
    def generate_excel(self,xml_path):
        tree = etree.parse(xml_path)  # 将xml解析为树结构
        root = tree.getroot()  # 获得该树的树根
        sheet_name = root.attrib["name"]
        case_id = 0 # 测试用例编号
        dict_list = []
        for child_root in root:
            case_title = ""  # 测试用例名称
            steps = ""  # 测试步骤
            expect_result = ""  # 期望结果
            priority = ""  # 优先级
            if child_root.tag == "testcase":
                # print(child_root.attrib)
                if "internalid" in child_root.attrib:
                    case_id = child_root.attrib["internalid"]
                else:
                    case_id += 1
                if "name" in child_root.attrib:
                    case_title = child_root.attrib["name"]
                # 查找优先级
                # iterfind(self, path, namespaces=None)
                # Iterates over all matching subelements, by tag name or path.

                for prio in child_root.iterfind("importance"):
                    # print(prio.text)
                    if prio.text == '3':
                        priority = "H"

                # 查找测试步骤和期望结果
                for i in child_root.iterfind("steps/step"):
                    # 查找测试步骤
                    for step in i.iterfind("actions"):
                        if step.text:
                            steps = step.text
                    # print(steps)
                    # print(">>>>>>>>>>>>>")

                    # 查找期望结果
                    for result in i.iterfind("expectedresults"):
                        if result.text:
                            expect_result += result.text.replace("\n","")

                    # 利用段落标签分割出测试步骤
                    steps = self.remove_html_char(steps)
                    steps = self.remove_space(steps)
                    step_item = re.split("\d\.|\d、", steps)  # 测试步骤以数字.开始或者数字、进行分割
                    step_content = ""
                    index = 1  # 步骤计数 +
                    # print(step_item)
                    for step in step_item:
                        if self.remove_space(step) != "":
                            step_content += str(index) + "." + step + "\n"
                            index += 1
                    steps = self.remove_space(case_title) + "\n" + "测试步骤\n" + step_content
                    # print(steps)
                    # print(">>>>>>>>>>>>>")

                    expect_result = self.remove_html_char(expect_result)
                    expect_result = self.remove_space(expect_result+"\n")
        #             testcase_dict = {"testcase_id": case_id, "testcase": steps, "expected_result": expect_result,"priority": priority}
        #             dict_list.append(testcase_dict)
        # dict_list.append({"sheet_name": sheet_name})
        # return dict_list

                    # 使用yield
                    yield {"testcase_id": case_id, "testcase": steps, "expected_result": expect_result,
                                     "priority": priority}
                yield {"sheet_name": sheet_name}

    # 去除xml文件中存在的html标签
    def remove_html_char(self,line):
        if line:
            result = line
            for char in HTML_CHAR:
                result = result.replace(char,'')
            return result
        else:
            return ""

    # 去除空格和空行
    def remove_space(self,line):
        if line:
            result = line
            result = result.replace(" ","").replace("\t","").replace("\n","")
            return result
        else:
            return ""

    # 设置单元格样式
    def set_style(self,cell,style):
        for key,value in style.items():
            if key == "alignment": # 单元格的对齐方式
                cell.alignment = value
            if key == "font": # 单元格字体
                cell.font = value
            if key == "border": # 单元格的边框
                cell.border = value
            if key == "protection": # 单元格写保护
                cell.protection = value


    def wriet2Excel(self,path):
        """
        :param path: 保存路径
        :return:
        """
        # 将模板中的内容复制到新建的excel文件
        shutil.copyfile(os.path.abspath("template.xlsx"),path)
        wb = openpyxl.load_workbook(path) # 打开excel文件
        sheet = wb.active
        row = 17
        # 设置单元格中文字自动换行
        alignment = Alignment(horizontal='general', vertical='bottom', text_rotation=0,
                              wrap_text=True,shrink_to_fit=True, indent=0)
        font = Font(name="宋体", size=10)  # 设置字体和大小
        style = {"alignment": alignment, "font": font}
        dict_list = self.generate_excel("demo.xml")
        for idx,val in enumerate(dict_list):
            col = 2
            for key,key_val in val.items():
                # 写入数据
                sheet.cell(row, col, key_val)
                self.set_style(sheet.cell,style)
                col += 1

            row += 1
        # sheet.title = dict_list[len(dict_list)-1]["sheet_name"]
        wb.save(os.path.join(path))  # 保存文件

    def wriet2Excel2(self,path,content):
        """
        :param content:转换后的xml内容
        :param path: 保存路径
        :return:
        """
        # 将模板中的内容复制到新建的excel文件
        shutil.copyfile(os.path.abspath(r"C:\Users\yangwei.li\PycharmProjects\ShowMeTheCode\xml2excel\template.xlsx"),path)
        wb = openpyxl.load_workbook(path) # 打开excel文件
        sheet = wb.active
        # 设置单元格中文字自动换行
        alignment = Alignment(horizontal='general', vertical='bottom', text_rotation=0,
                              wrap_text=True,shrink_to_fit=True, indent=0)
        font = Font(name="宋体", size=10)  # 设置字体和大小
        style = {"alignment": alignment, "font": font}
        # 将用例写入excel
        # content 是一个字典集合
        row = 17
        for line in content:
            if "sheet_name" not in line.keys():
                # 写入ID
                cell = sheet.cell(row, 2)
                cell.value = int(line['testcase_id'])
                self.set_style(cell, style)

                # 写入测试标题和步骤
                cell = sheet.cell(row, 3)
                cell.value = line['testcase']
                self.set_style(cell, style)

                # 写入期望结果
                cell = sheet.cell(row, 4)
                cell.value = line['expected_result']
                self.set_style(cell, style)

                # 写入优先级
                cell = sheet.cell(row, 5)
                cell.value = line['priority']
                self.set_style(cell, style)

                row += 1
            else:
                sheet.title = line["sheet_name"]  # 修改sheet名字为xml中test suite名

        wb.save(os.path.join(path))  # 保存文件


if __name__ == "__main__":
    demo = Xml2Excel()
    content = demo.generate_excel("C:/Users/yangwei.li/PycharmProjects/ShowMeTheCode/xml2excel/demo.xml")
    # demo.wriet2Excel("demo.xlsx")
    demo.wriet2Excel2("C:/Users/yangwei.li/PycharmProjects/ShowMeTheCode/xml2excel/demo3.xlsx",content)
    print("done......")

