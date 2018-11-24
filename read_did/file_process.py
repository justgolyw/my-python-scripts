#!/usr/bin/env python
# -*- coding:utf-8 -*-
import os,json
from datetime import datetime

import openpyxl


# 获取路径下指定后缀名的文件
def get_file_list(file_path, file_suffix):
    """
    :param file_path: 文件路径
    :param file_suffix: 文件后缀名
    :return: 返回一个list
    """

    file_list = []
    # 对文件的操作用try except包围
    try:
        # 1.路径为文件
        if os.path.exists(file_path) and os.path.isfile(file_path) \
                and os.path.splitext(file_path)[1] in file_suffix:
            file_list.append(file_path)
        # 2.路径为文件夹
        elif os.path.exists(file_path) and os.path.isdir(file_path):
            # 获取文件夹下的所有文件
            file_all = os.listdir(file_path)
            for each in file_all:
                cur_file = os.path.join(file_path, each)
                if os.path.isfile(cur_file) and os.path.splitext(cur_file)[1] in file_suffix:
                    file_list.append(cur_file)
        else:
            print('The file_path is wrong!!')

        if len(file_list) == 0:
            print('no file !!')
        else:
            return file_list
    except Exception as e:
        print(e)

# 为了让新输出的文件不会覆盖之前输出的文件
# 我们需要处理输出文件重名的问题
def save_file_rename(path_out,file_name):
    """
    :param path_out: 输出文件的保存路径
    :param file_name: 文件名字
    :return:
    """
    path_real = os.path.join(path_out,'result')
    if not os.path.exists(path_real):
        os.mkdir(path_real)
    save_path = os.path.join(path_real,'%s.xlsx' % file_name)

    if os.path.exists(save_path):
        os.rename(save_path,os.path.join(path_real,'%s_%s.xlsx' % (file_name,datetime.now().strftime("%Y%m%d_%H%M%S"))))

    return save_path

# 将DID的结果保存到excel
def write_to_excel_did(dict_file, path_out):
    """
    :param dict_file: 要保存的内容
    :param path_out: 保存路径
    :return:
    """
    # 创建文件夹
    save_path = save_file_rename(path_out,'DID_result')
    # dirname = os.path.join(path_out, 'did_result')
    # if not os.path.exists(dirname):
    #     os.mkdir(dirname)
    # save_file_name = 'DID_result.xlsx'  # excel文件的名字
    # save_path = os.path.join(dirname, save_file_name)
    wb = openpyxl.Workbook()
    # sheet = wb.active
    # sheet.title = 'DID_result'
    sheet_title = ['DID', 'Size', 'Analysis_data', 'Description', 'Analysis_result', 'Respect_result', 'Judge']
    sheet_index = 0
    # for file_key in sorted(dict_file.keys()):
    for file_key,file_value in dict_file.items():

        # file_key = unicode(file_key,'utf-8')
        # sheet = wb.create_sheet(file_key.decode("gb2312"),sheet_index)
        sheet = wb.create_sheet(file_key, sheet_index)
        sheet_index += 1

        for index, val in enumerate(sheet_title):
            sheet.cell(1, index + 1, val)
        index_row = 2  # 从第二行开始写入内容
        # index_col = 1
        # res_out = dict_file[file_key]
        res_out = file_value
        for key, value in res_out.items():
            print(key, value)
            for index, value2 in enumerate(value):
                # print(value2)
                sheet.cell(index_row, index + 1, value2)
            index_row += 1

    print('写入excel成功')
    wb.save(save_path)


# 分析can_trace日志
def logfile_trace_did(trace_path, find_msg):
    """
    :param trace_path:can_trace 路径
    :param find_msg: 要寻找的信息 分为2段："07A4;62" 对应于"提取的can_id;结果开始的值"
    :return:
    """
    # 读取json文件
    get_path = os.getcwd()  # 当前所在的工作路径
    # print("path",get_path)
    # while get_path.find('read_did') != -1:  # 回退到read_did的上一层
    #     get_path = os.path.abspath(os.path.join(get_path, '..'))
    json_did_path = os.path.abspath(os.path.join(get_path, r'did_nl3b.json'))

    with open(json_did_path, 'rb') as f:
        dict_did = json.load(f) # 读取json文件
        # result = json.dumps(dict_did, ensure_ascii=False)
        # print(type(dict_did))

    find_list = str(find_msg).split(';')  # str 转为 list
    path_out = os.path.dirname(trace_path)  # 文件的保存路径

    # 获取文件列表
    file_suffix = ['.trc','.txt','.asc']
    file_list = get_file_list(trace_path,file_suffix)
    dict_file = {}
    for index, each in enumerate(file_list):
        res_out = {}  # 保存最终的结果
        dic_out = {}
        cnt = 0
        cur_did = ''
        # 打开文件
        with open(each, 'r') as f:
            for line in f:  # 按行读取
                # print(line.decode("gb2312"))
                if str(line).find(find_list[0]) > -1:  # 定位到07A4
                    line_list = str(line).strip().split(' ')
                    while '' in line_list:
                        line_list.remove('')  # 去除中间的空格
                    # print(line_list)
                    if find_list[1] in line_list:  # 定位到62
                        index = line_list.index(find_list[1])
                        cur_did = '{0}{1}'.format(line_list[index + 1], line_list[index + 2])  # 当前did的值
                        # print(cur_did)
                        cnt += 1
                        # dic_out[cnt] = [cur_did,line_list] # 字典以did序号作为键，以当前的did和line_list作为值
                        dic_out[cur_did] = [line_list]  # 字典以did值作为键值

                    else:  # 不包含62说明是一组
                        # dic_out[cnt].append(line_list)
                        dic_out[cur_did].append(line_list)

            print(dic_out)
            temp_list = []  # 保存DID信息
            for k in dic_out.keys():
                did_size = 0  # did长度
                # cur_did = dic_out[k][0]
                cur_did = k
                did_des = dict_did[cur_did][1]  # DID描述
                flag_ascii = dict_did[cur_did][2]  # ASCII/无符号(默认ASCII)
                respect = dict_did[cur_did][3]  # 预期值
                for cur_index, cur_val in enumerate(dic_out[k]):
                    if cur_index == 0:
                        flage_index = cur_val.index(find_list[1])  # 从62开始才是有用的内容
                        did_size = int(cur_val[flage_index - 1], 16)  # DID长度
                        temp_list = cur_val[flage_index + 3:]
                    if cur_index > 0:
                        temp_list += cur_val[6:]
                # print(temp_list)
                # 转成str
                temp_str = ','.join(temp_list)
                # print(temp_str)
                # 转ASCII码
                if flag_ascii == 'True':
                    res_str = ''.join(chr(int(str(n), 16)) for n in temp_list)  # DID值
                else:
                    res_str = ''.join(str(n) for n in temp_list)
                if respect == res_str:
                    res_judge = 'True'
                else:
                    res_judge = 'False'

                # print(res_str)
                res_out[k] = [cur_did, did_size, temp_str, did_des, res_str, respect, res_judge]
            dict_file[os.path.basename(each)] = res_out
    write_to_excel_did(dict_file, path_out)


if __name__ == '__main__':
    file_path = r'22服务读取所有DID.asc'  # python3 对于中文路径能够自动转unicode
    # file_path = r'WM.trc'  # python3 对于中文路径能够自动转unicode
    # file_path = r'C:\Users\yangwei.li\Desktop\wmTRC'
    # 处理含有中文路径的文件
    # file_path2 = 'C:\Users\yangwei.li\Desktop' + u'\WM版本号.trc'
    # utf-8 转 unicode
    # file_path3 = unicode('C:\Users\yangwei.li\Desktop\WM版本号.trc', 'utf-8')
    find_msg = u'7A1;62'
    # find_msg = u'07A4;62'
    logfile_trace_did(file_path, find_msg)