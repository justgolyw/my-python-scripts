#!/usr/bin/env python
# -*- coding:utf-8 -*-
"""
第 0017 题： 将 第 0014 题中的 student.xls 文件中的内容写到 student.xml 文件中
"""
#!/usr/bin/env python
#coding=utf-8

import xlrd,json
import openpyxl
from lxml import etree
from xml.etree.ElementTree import Element,SubElement,Comment,ElementTree

from collections import OrderedDict

def xls2xml(xls_name):
    with xlrd.open_workbook(xls_name) as wb:
        ws = wb.sheet_by_index(0)
    table = OrderedDict()
    for i in range(ws.nrows):

        key = int(ws.row_values(i)[0])
        value = str(ws.row_values(i)[1:])
        table[key] = value
    print(table)

    with open("student.xml",'w',encoding="utf-8") as f:
        root = etree.Element("root")  # 父节点
        e_root = etree.ElementTree(root)
        child = etree.SubElement(root,'students') # 子节点
        # 添加注释
        child.append(etree.Comment('\n    studet_info\n    "id" : [name,math,Chinese,English]\n'))
        # json.dumps: dict 转 str
        child.text='\n'+str(json.dumps(table, indent=4, ensure_ascii=False))+'\n'

        f.write('<?xml version="1.0" encoding="UTF-8"?>\n')
        f.write(etree.tounicode(e_root.getroot()))

def xls2xml_2(xls_name):
    wb = openpyxl.load_workbook(xls_name)
    # ws = wb.active
    # ws = wb.get_active_sheet() # python3 中推荐使用 wb.active,wb.get_active_sheet()方法已过时
    # ws = wb.get_sheet_by_name("example") # python3 推荐使用 wb[sheetname]
    # wb.active = 1  # 指定sheet的索引
    ws = wb.active
    table = OrderedDict()
    for i in range(1,ws.max_row+1):
        val = []
        key = int(ws.cell(i,1).value)
        for j in range(2,ws.max_column+1):
            val.append(ws.cell(i, j).value)
        table[key] = str(val)
    # # jsonObj = json.dumps(table, indent=4, ensure_ascii=False)
    jsonObj = str(json.dumps(table,indent=4,ensure_ascii=False))

    print(table)
    print(jsonObj)

    root = Element("root") # 根节点
    comment = Comment('学生信息表\n"id" : [名字, 数学, 语文, 英文]\n') # 注释
    # root.append(comment) # 添加注释
    child = SubElement(root,'students') # 子节点
    child.append(comment)
    # root.append(comment)  # 添加注释
    child.text = jsonObj+"\n"
    tree = ElementTree(root) # 构造一棵树
    tree.write('student.xml', encoding='utf8')


def txt2xml(file_name):
    with open(file_name,'r',encoding="utf-8") as f:
        # 使用load函数将包含json数据的文本转化为python(dict)对象
        data = json.load(f)
        print(data)
        # 将dict 转成 str
        data_str = json.dumps(data,indent=4, ensure_ascii=False)
        print(data_str)



def xml2excel(file_name):
    with open(file_name,'r',encoding="utf-8") as f:
        text = f.read()
    root = etree.fromstring(text.encode("utf-8"))
    print(etree.tostring(root,pretty_print="true").decode("utf-8"))

    content = root[0].text
    print(content)
    jsObj = json.loads(content) # str 转 dict
    print(jsObj)
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.name = "example"
    index_col = 1
    for k,v in jsObj.items():
        _list = v.split(",")
        _list[0] = _list[0].replace('[', '')
        _list[len(_list) - 1] = _list[len(_list) - 1].replace(']', '')
        sheet.cell(index_col,1,k)
        for idx,value in enumerate(_list):
            sheet.cell(index_col,idx+2,value)
        index_col += 1
    wb.save("MyStudent.xlsx")

if __name__=='__main__':
    # xls2xml_2('student.xlsx')
    # txt2xml('city.txt')
    xml2excel('student.xml')


